#!/usr/bin/env python3
"""
Jim Daws Trucking - Driver Settlement PDF to QuickBooks Excel
Usage: python process_settlements.py <settlement.pdf> [output.xlsx]

PARSING FLOW:
  1. OTHER PAY/DEDUCTIONS section
  2. TOTAL SETTLEMENT section
  3. RESERVES section
"""

import sys
import re
from datetime import datetime, timedelta
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font
import streamlit as st
from streamlit_gsheets import GSheetsConnection
import pandas as pd


def load_drivers_from_sheets():
    try:
        # Establish connection using your Streamlit Secrets configuration
        conn = st.connection("gsheets", type=GSheetsConnection)
        df = conn.read(worksheet="Drivers", ttl="5m") # Caches for 5 minutes
        
        drivers_dict = {}
        for _, row in df.iterrows():
            code = str(row['Driver Code']).strip()
            
            # Reconstruct the expected nested structure
            notes_rec = ("", "")
            if pd.notna(row['Notes Rec Desc']) and str(row['Notes Rec Desc']).strip():
                notes_rec = (str(row['Notes Rec Desc']).strip(), str(row['Notes Rec Account']).strip())
                
            maint_fund = ("", "")
            if pd.notna(row['Maint Fund Desc']) and str(row['Maint Fund Desc']).strip():
                maint_fund = (str(row['Maint Fund Desc']).strip(), str(row['Maint Fund Account']).strip())
                
            drivers_dict[code] = (
                str(row['Driver Name']).strip(),
                str(row['Truck']).strip(),
                str(row['Class']).strip(),
                notes_rec,
                maint_fund
            )
        return drivers_dict
    except Exception as e:
        st.error(f"Failed to load driver mappings from Google Sheets: {e}")
        return {}

# Re-establish the global DRIVERS variable dynamically
# {driver code: (driver name, truck, class, notes receivable, maintenance fund)}
DRIVERS = load_drivers_from_sheets()

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_amount(text):
    text = re.sub(r'[,$]', '', text).strip()
    try:
        return float(text)
    except ValueError:
        return None

def new_settlement(settlement_date, driver_code, driver_name, driver_truck, driver_class):
    return {
        'settlement_date': settlement_date,
        'due_date': settlement_date + timedelta(days=1),
        'driver_code': driver_code,
        'driver_name': driver_name,
        'truck': driver_truck,
        'driver_class': driver_class,
        'invoice_number': None,
        'gross_pay': None,
        'total_trip_expenses': 0,
        'phys_dam_ins': None,
        'liab_cargo_ins': None,
        'occ_acc_ins': None,
        'truck_note': None,
        'maint_fund': None,
        'efs_mastercard': 0,
        'licensing_2025': None,
        'licensing_2026': None,
        'owner_operator_escrow': None,
        'loan': 0,
        'garnishments': [],
        'scales': [],
        'unmapped_other_pay': [],
        'unmapped_reserves': [],
    }

# ── PDF parsing ───────────────────────────────────────────────────────────────

SEC_NONE             = 0
SEC_OTHER_PAY        = 1
SEC_TOTAL_SETTLEMENT = 2
SEC_RESERVES         = 3

def extract_settlements(pdf_path):
    settlements = []
    current = None
    section = SEC_NONE
    settlement_date = None
    current_reserve_type = None
    current_licensing_year = None
    current_unknown_name = None

    all_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

    for line in all_lines:
        line = line.strip()
        if not line:
            continue

        # ── Settlement date from page header ──────────────────────────────────
        m = re.search(r'FINAL TRUCKING SETTLEMENTS as of (\d{2}/\d{2}/\d{4})', line)
        if m and settlement_date is None:
            settlement_date = datetime.strptime(m.group(1), '%m/%d/%Y')
            continue

        # ── New driver ─────────────────────────────────────────────────────────
        m = re.search(r'Payables:\s+.+?\s+\(([A-Z]+)\)', line)
        if m:
            code = m.group(1)
            if current is None or code != current['driver_code']:
                if current:
                    settlements.append(current)
                if code not in DRIVERS:
                    print(f"  ⚠ Unknown driver code: {code} — skipping")
                    current = None
                    section = SEC_NONE
                    continue
                name = DRIVERS.get(code, code)[0]
                truck = DRIVERS.get(code, code)[1]
                driver_class = DRIVERS.get(code, code)[2]
                current = new_settlement(settlement_date, code, name, truck, driver_class)
                section = SEC_NONE
                current_reserve_type = None
                current_licensing_year = None
                current_unknown_name = None
            continue

        if current is None:
            continue

        # ── Section detection ──────────────────────────────────────────────────
        if re.match(r'OTHER PAY/DEDUCTIONS', line) and 'TOTAL' not in line:
            section = SEC_OTHER_PAY
            continue

        if re.match(r'TOTAL SETTLEMENT', line):
            section = SEC_TOTAL_SETTLEMENT
            continue

        if re.match(r'RESERVES$', line):
            section = SEC_RESERVES
            continue

        # ═════════════════════════════════════════════════════════════════════
        # SECTION 1 — OTHER PAY/DEDUCTIONS
        # ═════════════════════════════════════════════════════════════════════
        if section == SEC_OTHER_PAY:
            
            if re.match(r'TOTAL OTHER PAY/DEDUCTIONS', line):
                section = SEC_NONE
                continue

            m = re.search(r'(-?[\$]?[\d,]+\.\d{2})\s*$', line)
            if m:
                amt = parse_amount(m.group(1))
            else:
                continue

            mapped = False

            # EFS Mastercard Advance 
            if re.search(r'^\s*(\d{2}/\d{2}/\d{2})\s+ADVANCE', line, re.I):
                current['efs_mastercard'] += amt
                mapped = True

            # Fuel expenses that weren't associated with a specific trip 
            elif re.search(r'^\s*\d{2}/\d{2}/\d{2}.*?Gals', line, re.I):
                current['total_trip_expenses'] += amt
                mapped = True
            
            # DEF expenses that weren't associated with a specific trip 
            elif re.search(r'^\s*\d{2}/\d{2}/\d{2}.*?\bDEF\b', line):
                current['total_trip_expenses'] += amt
                mapped = True

            # Physical Damage Ins
            elif re.search(r'PHYS(ICAL)?\s*DAM(AGE)?\s*INS', line, re.I):
                current['phys_dam_ins'] = amt
                mapped = True

            # Liability & Cargo Ins
            elif re.search(r'LIAB.*(CARGO|INS)', line, re.I):
                current['liab_cargo_ins'] = amt
                mapped = True

            # OCC ACC INS
            elif re.search(r'OCC\s*ACC\s*INS', line, re.I):
                current['occ_acc_ins'] = amt
                mapped = True

            # Truck Note / Payment
            elif re.search(r'TRUCK\s+(NOTE|PAYMENT)', line, re.I):
                current['truck_note'] = amt
                mapped = True

            # Child support / garnishments
            elif re.search(r'CHILD\s+SUPPORT', line, re.I):
                current['garnishments'].append(amt)
                mapped = True

            # Scales & Tolls
            elif re.match(r'^\d{2}/\d{2}/\d{2}', line) and re.search(r'SCALE|TOLL', line, re.I):
                current['scales'].append(amt)
                mapped = True

            # RESERVE lines
            elif line.startswith('RESERVE'):
                if re.search(r'2025\s*Licens', line, re.I):
                    current['licensing_2025'] = amt
                    mapped = True
                elif re.search(r'2026\s*Licens', line, re.I):
                    current['licensing_2026'] = amt
                    mapped = True
                elif re.search(r'Owner/Operator\s*Escrow', line, re.I):
                    current['owner_operator_escrow'] = amt
                    mapped = True
                elif re.search(r'Maint', line, re.I):
                    current['maint_fund'] = amt
                    mapped = True
                elif re.search(r'Loan', line, re.I):
                    current['loan'] += amt
                    mapped = True

            if not mapped:
                # Isolate everything before the trailing dollar value
                m = re.search(r'^(.*?)\s*(-?[\$]?[\d,]+\.\d{2})\s*$', line)
                
                # Clean up trailing spaces or ledger dots/dashes (e.g., "MISC PAY ----")
                description = m.group(1).rstrip(' .-').strip() if m else line
                current['unmapped_other_pay'].append((description, amt, 'OTHER PAY/DEDUCTIONS'))

        # ═════════════════════════════════════════════════════════════════════
        # SECTION 2 — TOTAL SETTLEMENT
        # ═════════════════════════════════════════════════════════════════════
        elif section == SEC_TOTAL_SETTLEMENT:

            if re.search(r'PERCENTAGE PAY', line, re.I):
                m = re.search(r'\$?([\d,]+\.\d{2})\s*$', line)
                if m:
                    val = float(m.group(1).replace(',', ''))
                    current['gross_pay'] = (current['gross_pay'] or 0) + val

            elif re.search(r'TOTAL TRIP EXPENSES', line, re.I):
                amounts = re.findall(r'-[\$]?([\d,]+\.\d{2})', line)
                if amounts:
                    current['total_trip_expenses'] += -float(amounts[-1].replace(',', ''))

            # ── AP Invoice number ──────────────────────────────────────────────────
            elif re.search(r'NET PAY \(AP invo#\s+(.+?)\)', line):
                m = re.search(r'NET PAY \(AP invo#\s+(.+?)\)', line)
                current['invoice_number'] = m.group(1)

        # ═════════════════════════════════════════════════════════════════════
        # SECTION 3 — RESERVES
        # ═════════════════════════════════════════════════════════════════════
        elif section == SEC_RESERVES:
            line = line.strip()
            if not line:
                continue

            # ── SECTION EXIT GUARD ───────────────────────────────────────────
            # If we hit the summary table header, shut down Section 3 immediately
            if re.search(r'This Settlement\s*Year-to-Date', line, re.I):
                section = SEC_NONE
                current_reserve_type = None
                current_licensing_year = None
                current_unknown_name = None
                continue

            # 1. Look for Ignore Keywords at the start of the line
            if re.match(r'^Addition to Reserve:', line, re.I):
                continue

            if re.match(r'^End Reserve Balance', line, re.I):
                current_licensing_year = None # Reset the year tracker
                current_reserve_type = None  # Close the active ledger window
                current_unknown_name = None
                continue

            # 2. Match a standard dollar amount at the end of the line
            m = re.search(r'^(.*?)\s*(-?[\$]?[\d,]+\.\d{2})\s*$', line)
            
            if m:
                description = m.group(1).rstrip(' .-').strip()
                amt = parse_amount(m.group(2)) * -1

                # 3. Detect when a new sub-ledger window opens (Initial Balance rows)
                # We check the description to see if it names a reserve type
                if re.search(r'Owner/Operator\s*Escrow', description, re.I):
                    current_reserve_type = 'ESCROW'
                    continue  # Ignore the initial balance value
                
                elif re.search(r'Licensing', description, re.I):
                    year_match = re.search(r'(\d{4})\s*Licensing', description, re.I)
                    
                    # Save the year to a separate tracking variable
                    current_licensing_year = year_match.group(1) if year_match else "UNKNOWN"
                    current_reserve_type = 'LICENSING' 
                    continue
                
                elif re.search(r'Maint', description, re.I):
                    current_reserve_type = 'MAINTENANCE'
                    continue  # Ignore the initial balance value
                
                elif re.search(r'Loan', description, re.I):
                    current_reserve_type = 'LOAN'
                    continue  # Ignore the initial balance value

                # If it didn't match known accounts, and no window is open, this IS an initial balance header!
                elif current_reserve_type is None:
                    current_unknown_name = description
                    current_reserve_type = 'UNKNOWN'
                    continue  # Ignore this initial balance value and loop to next line

                # 4. Target Window Capture 
                # If we are inside an active ledger and it didn't match the headers above, it's a transaction!
                if current_reserve_type == 'ESCROW':
                    # Append a clean structured tuple: (Account Type, Row Description, Value)
                    current['unmapped_reserves'].append((description, amt, 'RESERVE ESCROW'))

                elif current_reserve_type == 'LICENSING':
                    # Append a clean structured tuple: (Account Type, Row Description, Value)
                    current['unmapped_reserves'].append((description, amt, 'RESERVE LICENSING', current_licensing_year))

                elif current_reserve_type == 'MAINTENANCE':
                    # Append a clean structured tuple: (Account Type, Row Description, Value)
                    current['unmapped_reserves'].append((description, amt, 'RESERVE MAINTENANCE'))

                elif current_reserve_type == 'LOAN':
                    # Append a clean structured tuple: (Account Type, Row Description, Value)
                    current['unmapped_reserves'].append((description, amt, 'RESERVE LOAN'))

                elif current_reserve_type == 'UNKNOWN':
                    # Explicitly appends the dynamic name so you see exactly what it is in your Excel
                    current['unmapped_reserves'].append((current_unknown_name, amt, 'RESERVE UNKNOWN'))

    if current:
        settlements.append(current)

    return settlements

# ── Excel output ──────────────────────────────────────────────────────────────

HEADERS = [
    'Post?', 'Invoice/Bill Date', 'Due Date', 'Invoice / Bill Number',
    'Transaction Type', 'Customer', 'Vendor', 'Currency Code',
    'Product/Services', 'Description', 'Qty', 'Discount %',
    'Unit Price', 'Category', 'Location', 'Class'
]

def make_row(s, description, unit_price, category):
    return [
        'No',
        s['settlement_date'],
        s['due_date'],
        s['invoice_number'],
        'Bill',
        None,
        s['driver_name'],
        'USD',
        None,
        description,
        None,
        None,
        unit_price,
        category,
        None,
        s['driver_class'],
    ]

def settlement_to_rows(s):
    rows = []
    driver_code = s.get('driver_code', '') or ''

    # ── TOTAL SETTLEMENT section ───────────────────────────────────────────────
    if s['gross_pay'] is not None:
        rows.append(make_row(s, 'Settlements', s['gross_pay'],
                             '53400.3500 Driver Pay:I/C Settlements'))

    if s['total_trip_expenses'] != 0:
        rows.append(make_row(s, 'Diesel', s['total_trip_expenses'],
                             '54100.1500 Truck Expense:Diesel'))

    # ── OTHER PAY/DEDUCTIONS section ──────────────────────────────────────────
    if s['phys_dam_ins'] is not None:
        rows.append(make_row(s, 'Physical Damage Ins - Trucks', s['phys_dam_ins'],
                             '54580.1500 Truck Expense:Physical Damage Ins'))

    if s['liab_cargo_ins'] is not None:
        rows.append(make_row(s, 'Liability & Cargo Ins', s['liab_cargo_ins'],
                             '54520.1500 Truck Expense:Liability & Cargo Ins'))

    if s['occ_acc_ins'] is not None:
        rows.append(make_row(s, 'Occupational Accident Insurance', s['occ_acc_ins'],
                             '53710.3500 Occupational Accident Insurance'))

    if s['truck_note'] is not None:
        truck_note = DRIVERS.get(driver_code)[3]
        if truck_note:
            rows.append(make_row(s, truck_note[0], s['truck_note'], truck_note[1]))
        else:
            rows.append(make_row(s, f"N/R - Truck # {s['truck']}",
                                 s['truck_note'], f"UNMAPPED - Notes Receivable Truck {s['truck']}"))

    if s['maint_fund'] is not None:
        maint_fund = DRIVERS.get(driver_code)[4]
        if maint_fund:
            rows.append(make_row(s, maint_fund[0], s['maint_fund'], maint_fund[1]))
        else:
            rows.append(make_row(s, f"Dr Maint Funds - Tr # {s['truck']}",
                                 s['maint_fund'], f"UNMAPPED - Maint Fund Truck {s['truck']}"))
    if s['loan'] != 0:
            rows.append(make_row(s, 'Loan', s['loan'], '10051.9000 Truck Expense:Comdata Comcheck'))
             
    if s['efs_mastercard'] != 0:
        rows.append(make_row(s, 'EFS Mastercard Advance', s['efs_mastercard'], 
                             '54090.1500 Truck Expense:EFS Mastercard Transaction'))

    if s['licensing_2025'] is not None:
        rows.append(make_row(s, '2025 Licensing Accrual', s['licensing_2025'],
                             '22397.9000 2025 Licensing Accrual'))

    if s['licensing_2026'] is not None:
        rows.append(make_row(s, '2026 Licensing Accrual', s['licensing_2026'],
                             '22398.9000 2026 Licensing Accrual'))

    if s['owner_operator_escrow'] is not None:
        rows.append(make_row(s, 'Owner/Operator Escrow', s['owner_operator_escrow'],
                             '22418.9000 Owner/Operator Escrow'))

    if s['garnishments']:
        amt = sum(s['garnishments'])
        rows.append(make_row(s, 'Garnishments', amt, '21700.9000 Garnishments'))

    if s['scales']:
        amt = sum(s['scales'])
        rows.append(make_row(s, 'Scales', amt, '55500.1500 Truck Expense:Scales & Tolls'))
        
    for line in s['unmapped_other_pay']:
        description = line[0]
        amt = line[1]
        section = line[2]
        rows.append(make_row(s, description, amt, section))
    
    for line in s['unmapped_reserves']:
        description = line[0]
        amt = line[1]
        reserve_type = line[2]

        if reserve_type == 'RESERVE ESCROW':
            rows.append(make_row(s, description, amt, '22418.9000 Owner/Operator Escrow'))

        elif reserve_type == 'RESERVE LICENSING':
            year = line[3]
            if year == '2025':
                rows.append(make_row(s, description, amt, '22397.9000 2025 Licensing Accrual'))
            elif year == '2026':
                rows.append(make_row(s, description, amt, '22398.9000 2026 Licensing Accrual'))
            
        elif reserve_type == 'RESERVE MAINTENANCE':
            driver_maint_fund = DRIVERS.get(s['driver_code'])[4]
            rows.append(make_row(s, driver_maint_fund[0], amt, driver_maint_fund[1]))

        elif reserve_type == 'RESERVE LOAN':
            rows.append(make_row(s, description, amt, '10051.9000 Truck Expense:Comdata Comcheck'))
        
        elif reserve_type == 'RESERVE UNKNOWN':
            rows.append(make_row(s, description, amt, reserve_type))

    rows.append(make_row(s, '', '', ''))

    return rows

def write_excel(settlements, output_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: QB Import data ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Settlements'
    ws.append(HEADERS)
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = white_font

    for s in settlements:
        for row in settlement_to_rows(s):
            ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'MM/DD/YYYY'

    wb.save(output_path)

    total_rows = sum(len(settlement_to_rows(s)) for s in settlements)
    print(f"Saved: {output_path}")
    print(f"  {len(settlements)} driver settlements → {total_rows} QB rows")

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python process_settlements.py <settlement.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else pdf_path.replace('.pdf', '_QB.xlsx')

    print(f"Processing: {pdf_path}")
    settlements = extract_settlements(pdf_path)
    print(f"Found: {len(settlements)} driver settlements")

    for s in settlements:
        notes_rec = DRIVERS.get(s['driver_code'])[3]
        maint_fund = DRIVERS.get(s['driver_code'])[4]
        flag = " ⚠ UNMAPPED" if (
            (s['truck_note'] is not None and not notes_rec) or
            (s['maint_fund'] is not None and not maint_fund)
        ) else ""
        print(f"  {s['driver_name']:<30} INV:{s['invoice_number']:<16} "
              f"Gross:{s['gross_pay']}{flag}")

    write_excel(settlements, output_path)
